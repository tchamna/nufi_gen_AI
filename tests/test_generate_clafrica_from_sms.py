from importlib.util import module_from_spec, spec_from_file_location
from pathlib import Path

from openpyxl import Workbook


MODULE_PATH = (
    Path(__file__).resolve().parents[1]
    / "android-keyboard"
    / "tools"
    / "generate_clafrica_from_sms.py"
)


def load_generator_module():
    spec = spec_from_file_location("generate_clafrica_from_sms", MODULE_PATH)
    module = module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(module)
    return module


def test_normalize_sms_text_applies_dictionary_maps():
    module = load_generator_module()
    replacements = (
        ("\u0259\u0300'", "\u00e8'"),
        ("\u00ec", "i"),
        ("\u00e0", "a"),
    )

    assert module.normalize_sms_text("h\u0259\u0300' s\u00ec m\u00e0", replacements) == "h\u00e8' si ma"


def test_load_sms_mapping_normalizes_xls_replacement_text(tmp_path):
    module = load_generator_module()
    workbook_path = tmp_path / "sms.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = module.SMS_SHEET_NAME
    sheet.cell(row=1, column=2, value="shortcut")
    sheet.cell(row=1, column=3, value="h\u0259\u0300' s\u00ec m\u00e0")
    workbook.save(workbook_path)

    mapping, duplicate_keys, usable_rows = module.load_sms_mapping(
        workbook_path=workbook_path,
        replacements=(
            ("\u0259\u0300'", "\u00e8'"),
            ("\u00ec", "i"),
            ("\u00e0", "a"),
        ),
    )

    assert mapping == {"shortcut": "h\u00e8' si ma"}
    assert duplicate_keys == {}
    assert usable_rows == 1
