from __future__ import annotations

import json
import importlib.util
import re
import subprocess
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
REPO_ROOT = ROOT.parent
WORKBOOK_PATH = ROOT / "nufi_sms_and_calendar.xlsx"
DICTIONARY_MAPS_PATH = ROOT / "dictionary_maps.py"
BASE_MAPPING_PATH = REPO_ROOT / "clafricaMapping.ts"
ASSETS_DIR = ROOT / "app" / "src" / "main" / "assets"
CLAFRICA_OUTPUT_PATH = ASSETS_DIR / "clafrica.json"
SMS_OUTPUT_PATH = ASSETS_DIR / "nufi_sms.json"
CALENDAR_OUTPUT_PATH = ASSETS_DIR / "nufi_calendar.json"
SMS_SHEET_NAME = "Nufi_SMS"
CALENDAR_SHEET_NAME = "Nufi_Calendar"
NUMERIC_APOSTROPHE_SHORTCUT = re.compile(r"^(\d+)'$")


def load_base_mapping() -> dict[str, str]:
    if BASE_MAPPING_PATH.exists():
        result = subprocess.run(
            [
                "npx.cmd",
                "tsx",
                "-e",
                (
                    "import { Clafrica } from './clafricaMapping.ts'; "
                    "process.stdout.write(JSON.stringify(Clafrica));"
                ),
            ],
            cwd=REPO_ROOT,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            check=False,
        )
        if result.returncode != 0:
            stderr = result.stderr.decode("utf-8", errors="replace").strip()
            raise RuntimeError(f"Failed to load base Clafrica map: {stderr}")
        return json.loads(result.stdout.decode("utf-8"))

    raise FileNotFoundError(
        f"Could not find base mapping source {BASE_MAPPING_PATH}"
    )


def load_dictionary_replacements() -> tuple[tuple[str, str], ...]:
    if not DICTIONARY_MAPS_PATH.exists():
        raise FileNotFoundError(
            f"Could not find dictionary map source {DICTIONARY_MAPS_PATH}"
        )

    spec = importlib.util.spec_from_file_location(
        "android_keyboard_dictionary_maps",
        DICTIONARY_MAPS_PATH,
    )
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Failed to load dictionary maps from {DICTIONARY_MAPS_PATH}")

    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)

    replacements: dict[str, str] = {}
    for attr_name in ("vowel_bana_to_komako", "dict_ton_bas"):
        mapping = getattr(module, attr_name, None)
        if not isinstance(mapping, dict):
            raise RuntimeError(f"Expected {attr_name} to be a dict in {DICTIONARY_MAPS_PATH}")
        replacements.update(mapping)

    return tuple(sorted(replacements.items(), key=lambda item: len(item[0]), reverse=True))


def normalize_sms_text(text: str, replacements: tuple[tuple[str, str], ...]) -> str:
    normalized = text
    for source, target in replacements:
        normalized = normalized.replace(source, target)
    return normalized


def normalize_sms_shortcut(shortcut: str) -> str:
    match = NUMERIC_APOSTROPHE_SHORTCUT.fullmatch(shortcut)
    if match:
        return f"{match.group(1)}*"
    return shortcut


def load_sms_mapping(
    workbook_path: Path = WORKBOOK_PATH,
    replacements: tuple[tuple[str, str], ...] | None = None,
) -> tuple[dict[str, str], dict[str, list[int]], int]:
    workbook = load_workbook(workbook_path, data_only=True)
    sheet = workbook[SMS_SHEET_NAME]
    active_replacements = replacements if replacements is not None else load_dictionary_replacements()

    mapping: dict[str, str] = {}
    duplicates: dict[str, list[int]] = defaultdict(list)
    usable_rows = 0

    for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        shortcut = "" if row[1] is None else normalize_sms_shortcut(str(row[1]).strip())
        replacement = "" if row[2] is None else normalize_sms_text(
            str(row[2]).strip(),
            active_replacements,
        )
        if not shortcut or not replacement:
            continue

        usable_rows += 1
        if shortcut in mapping:
            duplicates[shortcut].append(row_number)
        else:
            duplicates[shortcut] = [row_number]
        mapping[shortcut] = replacement

    duplicate_keys = {key: rows for key, rows in duplicates.items() if len(rows) > 1}
    return mapping, duplicate_keys, usable_rows


def parse_calendar_date(date_text: str) -> datetime:
    return datetime.strptime(date_text, "%B %d %Y")


def load_calendar_mapping(
    workbook_path: Path = WORKBOOK_PATH,
    replacements: tuple[tuple[str, str], ...] | None = None,
) -> tuple[dict[str, str], dict[str, list[int]], int]:
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    sheet = workbook[CALENDAR_SHEET_NAME]
    active_replacements = replacements if replacements is not None else load_dictionary_replacements()

    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    header_index = {str(value).strip(): index for index, value in enumerate(header_row) if value is not None}
    date_index = header_index["Date"]
    day_eng_index = header_index["Day(Eng)"]
    day_nufi_index = header_index["Day(Nufi)"]

    mapping: dict[str, str] = {}
    duplicates: dict[str, list[int]] = defaultdict(list)
    usable_rows = 0

    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        date_text = "" if row[date_index] is None else str(row[date_index]).strip()
        day_eng = "" if row[day_eng_index] is None else str(row[day_eng_index]).strip()
        day_nufi = "" if row[day_nufi_index] is None else normalize_sms_text(
            str(row[day_nufi_index]).strip(),
            active_replacements,
        )
        if not date_text or not day_eng or not day_nufi:
            continue

        parsed_date = parse_calendar_date(date_text)
        canonical_date = parsed_date.strftime("%d-%m-%Y")
        output = normalize_sms_text(
            (
                f"Zě'é mɑ́ {day_nufi} ({day_eng}), "
                f"līē' {parsed_date.strftime('%d')}, ngu' {parsed_date.strftime('%Y')}"
            ),
            active_replacements,
        )

        usable_rows += 1
        if canonical_date in mapping:
            duplicates[canonical_date].append(row_number)
        else:
            duplicates[canonical_date] = [row_number]
        mapping[canonical_date] = output

    duplicate_keys = {key: rows for key, rows in duplicates.items() if len(rows) > 1}
    return mapping, duplicate_keys, usable_rows


def main() -> None:
    replacements = load_dictionary_replacements()
    base_mapping = load_base_mapping()
    sms_mapping, duplicate_keys, usable_rows = load_sms_mapping(replacements=replacements)
    calendar_mapping, calendar_duplicate_keys, calendar_usable_rows = load_calendar_mapping(
        replacements=replacements,
    )
    overlap = sorted(set(base_mapping) & set(sms_mapping))
    CLAFRICA_OUTPUT_PATH.write_text(
        json.dumps(base_mapping, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )
    SMS_OUTPUT_PATH.write_text(
        json.dumps(sms_mapping, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )
    CALENDAR_OUTPUT_PATH.write_text(
        json.dumps(calendar_mapping, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )

    print(
        f"Wrote {len(base_mapping)} base shortcuts to {CLAFRICA_OUTPUT_PATH.name}"
    )
    print(f"Wrote {len(sms_mapping)} SMS shortcuts to {SMS_OUTPUT_PATH.name}")
    print(f"Wrote {len(calendar_mapping)} calendar dates to {CALENDAR_OUTPUT_PATH.name}")
    print(f"SMS usable rows: {usable_rows}")
    print(f"Calendar usable rows: {calendar_usable_rows}")
    print(f"Overlapping keys overridden by SMS: {len(overlap)}")
    if duplicate_keys:
        print("Duplicate shortcuts resolved by keeping the last row:")
        for key in sorted(duplicate_keys):
            print(f"  {key} -> rows {duplicate_keys[key]}")
    if calendar_duplicate_keys:
        print("Duplicate calendar dates resolved by keeping the last row:")
        for key in sorted(calendar_duplicate_keys):
            print(f"  {key} -> rows {calendar_duplicate_keys[key]}")


if __name__ == "__main__":
    main()
